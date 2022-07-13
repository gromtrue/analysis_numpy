import openpyxl
from time import sleep

# подключаемся к файлу
excel_file = openpyxl.load_workbook(r'C:\Users\gribanov_r\Desktop\2022.06.21 17.43 deliveries_today.xlsx')
# print(excel_file.sheetnames)

# tdsheet_sheet = excel_file['TDSheet']
# Подключаемся к активной вкладке
tdsheet_sheet = excel_file.active


# cell_obj = tdsheet_sheet.cell(row=1, column=1)
# print(type(cell_obj))
# print(f'Employees[A1]={cell_obj.value}')
# print(f'Employees[A1]={tdsheet_sheet["A1"].value}')

# Выводим количество строк и столбцов
# print(f'Всего строк = {tdsheet_sheet.max_row} и Всего Столбцов = {tdsheet_sheet.max_column}')


# header_cells_generator = tdsheet_sheet.iter_rows(max_row=1)
#
# for header_cells_tuple in header_cells_generator:
#     for i in range(len(header_cells_tuple)):
#         print(i)
        # print(header_cells_tuple[i], header_cells_tuple[i].value)
        # print(header_cells_tuple[i])


# Печатаем все строки
# for x in range(5, tdsheet_sheet.max_row):
#     print(tdsheet_sheet.cell(row=x, column=4).value)

# печатаем все столбцы
# for x in range(1, tdsheet_sheet.max_column+1):
#     print(tdsheet_sheet.cell(row=5, column=x).value)


count = 0
# Просматриваем все строки
for x in range(1, tdsheet_sheet.max_row):
    # В 4-й строке ищем только значение "Нет"
    if tdsheet_sheet.cell(row=x, column=4).value == 'Нет':
        # print(f'Строка {x} {tdsheet_sheet.cell(row=x, column=4).value}')
        text = []
        # Просматриваем все столбцы
        for y in range(1, tdsheet_sheet.max_column+1):
            text.append(tdsheet_sheet.cell(row=x, column=y).value)
            # print(tdsheet_sheet.cell(row=x, column=y).value)
        count += 1
        print(text)
print(count)
