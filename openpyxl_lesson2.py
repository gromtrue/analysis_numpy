import openpyxl
from time import sleep

# подключаемся к файлу
excel_file = openpyxl.load_workbook(r'C:\Users\gribanov_r\Desktop\2022.06.21 17.43 deliveries_today.xlsx')

# Подключаемся к активной вкладке
tdsheet_sheet = excel_file.active

count = 0  # счётчик
# Просматриваем все строки
number_post = []
for x in range(1, tdsheet_sheet.max_row):
    # В 4-й строке ищем только значение "Нет"
    if tdsheet_sheet.cell(row=x, column=4).value == 'Нет':
        # Выбираем только уникальные значения по 5-му столбцу
        if number_post != tdsheet_sheet.cell(row=x, column=5).value:
            number_post = tdsheet_sheet.cell(row=x, column=5).value
            text = []
            # Просматриваем все столбцы
            # for y in range(1, tdsheet_sheet.max_column + 1):
            number_column = [1, 4, 5, 7, 8]  # Перечисляем номера нужных столбцов
            for y in range(len(number_column)):
                text.append(tdsheet_sheet.cell(row=x, column=number_column[y]).value)
            count += 1
            print(text)

print(count)

